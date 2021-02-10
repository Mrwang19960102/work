# -*- coding: utf-8 -*-
# @File:       |   deal_excel.py 
# @Date:       |   2021/1/13 21:47
# @Author:     |   ThinkPad
# @Desc:       |  
import os
import xlrd
import openpyxl
import pandas as pd
import xlsxwriter
from deal_pdf import PDFConverter

f_excel_path = './excel文件夹'
f_pdf_path = './pdf文件夹'

if not os.path.exists(f_excel_path):
    print('创建excel文件夹')
    os.mkdir(f_excel_path)

if not os.path.exists(f_pdf_path):
    print('创建pdf文件夹')
    os.mkdir(f_pdf_path)


def trans_data():
    df = pd.read_excel('./原始信息.xlsx')
    for index, info in df.iterrows():
        em_num = info['员工编号']
        if em_num:
            em_num = str(em_num)
        else:
            em_num = None

        name = info['姓名']
        if name:
            pass
        else:
            name = None

        sex = info['性别']
        if sex:
            pass
        else:
            sex = None

        age = info['年龄']
        if age:
            age = str(age)
        else:
            age = None

        birth_date = info['出生日期']
        if birth_date:
            pass
        else:
            birth_date = str(birth_date)

        phone = info['联系电话']
        if phone:
            phone = str(phone)
        else:
            phone = None
        id = info['身份证号']
        if id:
            id = str(id)
        else:
            id = None

        p_landscape = info['政治面貌']
        if p_landscape:
            pass
        else:
            p_landscape = None

        department = info['所在部门']
        if department:
            pass
        else:
            department = None

        position = info['职务']
        if position:
            pass
        else:
            position = None

        induction_date = info['入职时间']
        if induction_date:
            induction_date = str(induction_date)
        else:
            induction_date = None

        address = info['家庭住址']
        if address:
            pass
        else:
            address = None

        community = info['所属社区']
        if community:
            pass
        else:
            community = None

        # 创建文件
        file_name = './{}/{}{}.xlsx'.format(f_excel_path, em_num, name)
        workbook = xlsxwriter.Workbook(file_name)
        merge_format = workbook.add_format({
            'bold': True,
            'font_size': 20,
            'align': 'center',  # 水平居中
            'valign': 'vcenter',  # 垂直居中
        })
        worksheet = workbook.add_worksheet()
        # 写入标题数据
        worksheet.merge_range('A6:I6', '疫情防控承诺函', merge_format)
        workbook.close()
        value_dict = {
            '所在部门：': {'row': '9', 'columns': '1', 'value': department},
            '姓    名：': {'row': '11', 'columns': '1', 'value': name},
            '性    别：': {'row': '13', 'columns': '1', 'value': sex},
            '身份证号：': {'row': '15', 'columns': '1', 'value': id},
            '联系电话：': {'row': '17', 'columns': '1', 'value': phone},
            '入职时间：': {'row': '19', 'columns': '1', 'value': induction_date},
            '家庭住址：': {'row': '21', 'columns': '1', 'value': address},
            '年    龄：': {'row': '13', 'columns': '3', 'value': age},
            '职    务：': {'row': '9', 'columns': '6', 'value': position},
            '员工编号：': {'row': '11', 'columns': '6', 'value': em_num},
            '出生日期：': {'row': '13', 'columns': '6', 'value': birth_date},
            '政治面貌：': {'row': '15', 'columns': '6', 'value': p_landscape},
            '所属社区：': {'row': '21', 'columns': '6', 'value': community},
        }

        wb = openpyxl.load_workbook(file_name)
        ws = wb.active
        ws.title = str(em_num) + str(name)
        for k, v in value_dict.items():
            ws.cell(row=int(v['row']), column=int(v['columns']), value=k)
            ws.cell(row=int(v['row']), column=int(v['columns']) + 1, value=v['value'])
        # sheet['A9'] = '所在部门：'
        # sheet['B9'] = department
        # sheet['A11'] = '姓    名：'
        # sheet['B11'] = name
        # sheet['A13'] = '性    别：'
        # sheet['B13'] = sex
        # sheet['A15'] = '身份证号：'
        # sheet['B15'] = str(id)
        # sheet['A17'] = '联系电话：'
        # sheet['B17'] = phone
        # sheet['A19'] = '入职时间：'
        # sheet['B19'] = induction_date
        # sheet['A21'] = '家庭住址：'
        # sheet['B21'] = address
        # sheet['D13'] = '年    龄：'
        # sheet['E13'] = str(age)
        # sheet['F9'] = '职    务：'
        # sheet['G9'] = position
        # sheet['F11'] = '员工编号：'
        # sheet['G11'] = str(em_num)
        # sheet['F13'] = '出生日期：'
        # sheet['G13'] = birth_date
        # sheet['F15'] = '政治面貌：'
        # sheet['G15'] = p_landscape
        # sheet['F21'] = '所属社区：'
        # sheet['G21'] = community
        ws['A24'] = '本人承诺：xxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'

        wb.save(file_name)
        print('name={}数据存储完成'.format(name))
        # 获取到文件的sheet数
        b = xlrd.open_workbook(file_name)
        sheetnum = len(b.sheets())
        pdfConverter = PDFConverter(file_name, sheetnum)
        pdfConverter.run_conver()


if __name__ == '__main__':
    trans_data()
